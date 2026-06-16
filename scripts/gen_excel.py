# -*- coding: utf-8 -*-
"""Generate the HuoMa Alibaba Cloud product list as a formatted Excel file."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "产品清单"

# ── Styles ──
hfont = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
bfont = Font(name="微软雅黑", size=10)
bfont_red = Font(name="微软雅黑", size=10, bold=True, color="C00000")
tfont = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
sheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sheader_font = Font(name="微软雅黑", bold=True, size=10)
total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
total_font = Font(name="微软雅黑", bold=True, size=11)
total_font_red = Font(name="微软雅黑", bold=True, size=12, color="C00000")
note_font = Font(name="微软雅黑", size=9, color="555555")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Column widths ──
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 14

# ── Title ──
ws.merge_cells("A1:D1")
c = ws["A1"]; c.value = "火马（HuoMa）阿里云部署 — 产品清单"; c.font = tfont; c.alignment = center
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:D2")
c = ws["A2"]; c.value = "2026-06-16 · 预估月费合计 ¥1,126/月"; c.font = Font(name="微软雅黑", size=9, color="666666"); c.alignment = center
ws.row_dimensions[2].height = 22

# ── Headers ──
for col, h in enumerate(["序号", "产品", "规格", "月费"], 1):
    c = ws.cell(row=3, column=col, value=h); c.font = hfont; c.fill = hfill; c.alignment = center; c.border = border
ws.row_dimensions[3].height = 24

# ── Data rows ──
products = [
    ["1", "ECS 应用服务器",       "4 vCPU · 8 GiB · AMD (c7a.xlarge) · Ubuntu 22.04 · 40G 系统盘",                  "¥240"],
    ["2", "弹性公网 IP (EIP)",     "按流量计费 · 50 Mbps 峰值 · BGP 多线",                                           "¥15"],
    ["3", "RDS MySQL 主节点",      "2 vCPU · 4 GiB · 高可用版 · 100G SSD · MySQL 8.0",                               "¥460"],
    ["4", "RDS MySQL 只读",       "1 vCPU · 2 GiB · 管理后台报表查询 · 共享主实例存储",                               "¥210"],
    ["5", "Redis 缓存",           "开源版 · 经典部署 · 2 GiB 标准版 (主从) · Redis 5.0",                              "¥198"],
    ["6", "OSS 对象存储",          "按量付费 · ~20 GiB · JAR包 / 配置 / 数据库备份 · 生命周期 30 天",                  "¥3"],
    ["7", "SSL 证书",             "免费 DV 单域名 · 云盾签发 · Nginx 格式",                                          "¥0"],
]

for i, row in enumerate(products):
    r = 4 + i
    for col, val in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=val); c.font = bfont; c.border = border
        if col == 1: c.alignment = center
        elif col == 4: c.alignment = center; c.font = bfont_red
        else: c.alignment = wrap
    ws.row_dimensions[r].height = 28

# ── Total row ──
tr = 4 + len(products)
for col in range(1, 5):
    c = ws.cell(row=tr, column=col); c.border = border; c.font = total_font; c.fill = total_fill
ws.merge_cells(f"A{tr}:C{tr}")
ws.cell(row=tr, column=1, value="合计"); ws.cell(row=tr, column=1).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=tr, column=4, value="¥1,126/月"); ws.cell(row=tr, column=4).alignment = center; ws.cell(row=tr, column=4).font = total_font_red
ws.row_dimensions[tr].height = 28

# ── Business Scenario ──
sr = tr + 2
ws.merge_cells(f"A{sr}:D{sr}")
for col in range(1, 5):
    c = ws.cell(row=sr, column=col); c.border = border; c.fill = sheader_fill
ws.cell(row=sr, column=1, value="业务场景"); ws.cell(row=sr, column=1).font = sheader_font; ws.cell(row=sr, column=1).alignment = center
ws.row_dimensions[sr].height = 24

dr = sr + 1
ws.merge_cells(f"A{dr}:D{dr}")
desc = (
    "书店为每所学校生成企业微信活码，家长扫码后系统自动分配服务老师、打标签（市/区/学校）、"
    "日限管控，满员自动从全局池轮换。日均承载最高 20 万次扫码，峰值瞬时 5000 人并发，"
    "数据存储于 RDS MySQL，Redis Stream 异步削峰消费回调事件。冷备 ECS 10 分钟内恢复。"
)
ws.cell(row=dr, column=1, value=desc).font = Font(name="微软雅黑", size=10)
ws.cell(row=dr, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
for col in range(1, 5):
    ws.cell(row=dr, column=col).border = border
ws.row_dimensions[dr].height = 55

# ── Notes ──
nr = dr + 2
ws.merge_cells(f"A{nr}:D{nr}")
for col in range(1, 5):
    c = ws.cell(row=nr, column=col); c.border = border; c.fill = sheader_fill
ws.cell(row=nr, column=1, value="备注"); ws.cell(row=nr, column=1).font = sheader_font; ws.cell(row=nr, column=1).alignment = center

notes = [
    "· ECS + RDS + Redis 均选包年，比按月付费节省 15-25%",
    "· 流量费按 20 万扫码/天估算约 ¥44/月，EIP 合计约 ¥60/月，上表按保守口径",
    "· 旧腾讯云服务器改为 Staging 测试机，兼做应急备机，不需额外买冷备 ECS",
    "· 域名需 ICP 备案通过后才能使用 443 端口接入企微回调",
]
for j, n in enumerate(notes):
    r = nr + 1 + j
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(row=r, column=1, value=n).font = note_font
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20

# ── Save ──
path = "docs/火马阿里云部署-产品清单.xlsx"
wb.save(path)
print(f"OK -> {path}")
